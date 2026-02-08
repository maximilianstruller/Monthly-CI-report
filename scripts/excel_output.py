"""
Excel output module - writes the competitive intelligence report to an .xlsx file.

Uses openpyxl to create a formatted Excel workbook with multiple tabs.
"""

import os
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# Style constants
HEADER_FONT = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
SUBHEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
COLUMN_HEADER_FONT = Font(name="Calibri", size=11, bold=True)
BODY_FONT = Font(name="Calibri", size=11)
DARK_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
MEDIUM_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
HIGH_RISK_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
MEDIUM_RISK_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
LOW_RISK_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
WRAP_ALIGNMENT = Alignment(wrap_text=True, vertical="top")


def _set_column_widths(ws, widths):
    """Set column widths from a list."""
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width


def _write_title(ws, title, report_date, col_span=4):
    """Write a title row with dark background spanning multiple columns."""
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=col_span)
    cell = ws.cell(row=1, column=1, value=title)
    cell.font = HEADER_FONT
    cell.fill = DARK_FILL
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=col_span)
    date_cell = ws.cell(row=2, column=1, value=f"Report Date: {report_date}")
    date_cell.font = Font(name="Calibri", size=11, italic=True, color="FFFFFF")
    date_cell.fill = DARK_FILL

    # Blank row 3
    return 4  # Next available row


def _write_column_headers(ws, row, headers):
    """Write column headers with styling."""
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = COLUMN_HEADER_FONT
        cell.fill = MEDIUM_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 20
    return row + 1


def write_executive_briefing(ws, briefing_bullets, report_date):
    """Write executive briefing tab."""
    _set_column_widths(ws, [8, 80, 15])
    row = _write_title(ws, "Executive Briefing", report_date, col_span=3)
    row = _write_column_headers(ws, row, ["#", "Key Insight", "Impact"])

    for i, bullet in enumerate(briefing_bullets, 1):
        impact = bullet.get("impact", "")
        ws.cell(row=row, column=1, value=i).font = BODY_FONT
        ws.cell(row=row, column=2, value=bullet.get("insight", "")).font = BODY_FONT
        impact_cell = ws.cell(row=row, column=3, value=impact)
        impact_cell.font = BODY_FONT

        # Color-code impact
        if impact == "HIGH":
            impact_cell.fill = HIGH_RISK_FILL
        elif impact == "MEDIUM":
            impact_cell.fill = MEDIUM_RISK_FILL
        elif impact == "LOW":
            impact_cell.fill = LOW_RISK_FILL

        for col in range(1, 4):
            ws.cell(row=row, column=col).border = THIN_BORDER
            ws.cell(row=row, column=col).alignment = WRAP_ALIGNMENT

        row += 1


def write_competitor_analysis(ws, competitors_data, report_date):
    """Write competitor analysis tab."""
    _set_column_widths(ws, [18, 12, 50, 40, 40, 40])
    row = _write_title(ws, "Competitor Analysis", report_date, col_span=6)
    row = _write_column_headers(ws, row, [
        "Competitor", "Risk Level", "Key Updates",
        "Strategic Implication", "Recommended Response", "Sources"
    ])

    for comp in competitors_data:
        risk = comp.get("risk_level", "")
        ws.cell(row=row, column=1, value=comp.get("name", "")).font = Font(name="Calibri", size=11, bold=True)
        risk_cell = ws.cell(row=row, column=2, value=risk)
        risk_cell.font = BODY_FONT
        if risk == "HIGH":
            risk_cell.fill = HIGH_RISK_FILL
        elif risk == "MEDIUM":
            risk_cell.fill = MEDIUM_RISK_FILL
        elif risk == "LOW":
            risk_cell.fill = LOW_RISK_FILL

        ws.cell(row=row, column=3, value=comp.get("updates", "")).font = BODY_FONT
        ws.cell(row=row, column=4, value=comp.get("implication", "")).font = BODY_FONT
        ws.cell(row=row, column=5, value=comp.get("recommendation", "")).font = BODY_FONT
        ws.cell(row=row, column=6, value=comp.get("sources", "")).font = BODY_FONT

        for col in range(1, 7):
            ws.cell(row=row, column=col).border = THIN_BORDER
            ws.cell(row=row, column=col).alignment = WRAP_ALIGNMENT

        ws.row_dimensions[row].height = 80
        row += 1


def write_risk_dashboard(ws, risk_data, report_date):
    """Write risk dashboard tab."""
    _set_column_widths(ws, [14, 18, 50, 50])
    row = _write_title(ws, "Risk Dashboard", report_date, col_span=4)
    row = _write_column_headers(ws, row, [
        "Risk Level", "Competitor", "Summary", "Impact"
    ])

    for level in ["HIGH", "MEDIUM", "LOW"]:
        items = risk_data.get(level, [])
        fill = {"HIGH": HIGH_RISK_FILL, "MEDIUM": MEDIUM_RISK_FILL, "LOW": LOW_RISK_FILL}[level]

        for item in items:
            level_cell = ws.cell(row=row, column=1, value=level)
            level_cell.font = Font(name="Calibri", size=11, bold=True)
            level_cell.fill = fill

            ws.cell(row=row, column=2, value=item.get("competitor", "")).font = BODY_FONT
            ws.cell(row=row, column=3, value=item.get("summary", "")).font = BODY_FONT
            ws.cell(row=row, column=4, value=item.get("impact", "")).font = BODY_FONT

            for col in range(1, 5):
                ws.cell(row=row, column=col).border = THIN_BORDER
                ws.cell(row=row, column=col).alignment = WRAP_ALIGNMENT

            ws.row_dimensions[row].height = 40
            row += 1


def write_positioning_matrix(ws, matrix_data, report_date):
    """Write positioning matrix tab."""
    _set_column_widths(ws, [30, 18, 30, 40])
    row = _write_title(ws, "Positioning Matrix", report_date, col_span=4)
    row = _write_column_headers(ws, row, [
        "Dimension", "Remerge Rating", "Leader(s)", "Gaps / Opportunities"
    ])

    for dimension in matrix_data:
        rating = dimension.get("remerge_rating", "")
        ws.cell(row=row, column=1, value=dimension.get("dimension", "")).font = Font(name="Calibri", size=11, bold=True)

        rating_cell = ws.cell(row=row, column=2, value=rating)
        rating_cell.font = BODY_FONT
        if rating == "Strong":
            rating_cell.fill = LOW_RISK_FILL
        elif rating == "Moderate":
            rating_cell.fill = MEDIUM_RISK_FILL
        elif rating == "Developing":
            rating_cell.fill = HIGH_RISK_FILL

        ws.cell(row=row, column=3, value=dimension.get("leaders", "")).font = BODY_FONT
        ws.cell(row=row, column=4, value=dimension.get("gaps", "")).font = BODY_FONT

        for col in range(1, 5):
            ws.cell(row=row, column=col).border = THIN_BORDER
            ws.cell(row=row, column=col).alignment = WRAP_ALIGNMENT

        ws.row_dimensions[row].height = 40
        row += 1


def write_strategic_synthesis(ws, synthesis_data, report_date):
    """Write strategic synthesis tab."""
    _set_column_widths(ws, [8, 80])
    row = _write_title(ws, "Strategic Synthesis", report_date, col_span=2)

    sections = [
        ("CROSS-COMPETITOR THEMES", synthesis_data.get("themes", [])),
        ("WHITE-SPACE OPPORTUNITIES", synthesis_data.get("opportunities", [])),
        ("VULNERABILITY POINTS", synthesis_data.get("vulnerabilities", [])),
        ("PRIORITY RECOMMENDATIONS", synthesis_data.get("recommendations", [])),
    ]

    for section_title, items in sections:
        # Section header
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        header_cell = ws.cell(row=row, column=1, value=section_title)
        header_cell.font = SUBHEADER_FONT
        header_cell.fill = DARK_FILL
        ws.row_dimensions[row].height = 25
        row += 1

        for i, item in enumerate(items, 1):
            ws.cell(row=row, column=1, value=i).font = BODY_FONT
            ws.cell(row=row, column=2, value=item).font = BODY_FONT
            for col in range(1, 3):
                ws.cell(row=row, column=col).border = THIN_BORDER
                ws.cell(row=row, column=col).alignment = WRAP_ALIGNMENT
            ws.row_dimensions[row].height = 30
            row += 1

        row += 1  # Blank row between sections


def write_raw_data(ws, raw_text, report_date):
    """Write raw search data for archival."""
    _set_column_widths(ws, [150])
    row = _write_title(ws, "Raw Search Data", report_date, col_span=1)

    lines = raw_text.split("\n")
    for line in lines:
        ws.cell(row=row, column=1, value=line[:32000]).font = Font(name="Consolas", size=10)
        ws.cell(row=row, column=1).alignment = WRAP_ALIGNMENT
        row += 1


def write_full_report(report_data, output_dir=None):
    """
    Main function: creates a formatted Excel workbook with all report sections.

    Args:
        report_data: dict containing all parsed report sections
        output_dir: directory to save the file (defaults to ./reports/)

    Returns:
        Path to the generated .xlsx file
    """
    report_date = datetime.now().strftime("%Y-%m-%d")
    filename = f"remerge_competitive_intel_{report_date}.xlsx"

    if not output_dir:
        output_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), "reports")
    os.makedirs(output_dir, exist_ok=True)

    filepath = os.path.join(output_dir, filename)

    print("Creating Excel workbook...")
    wb = Workbook()

    # Create tabs (rename default sheet for the first one)
    ws_briefing = wb.active
    ws_briefing.title = "Executive Briefing"
    ws_competitors = wb.create_sheet("Competitor Analysis")
    ws_risk = wb.create_sheet("Risk Dashboard")
    ws_matrix = wb.create_sheet("Positioning Matrix")
    ws_synthesis = wb.create_sheet("Strategic Synthesis")
    ws_raw = wb.create_sheet("Raw Data")

    print("Writing report sections...")

    write_executive_briefing(
        ws_briefing, report_data.get("executive_briefing", []), report_date
    )
    print("  Written: Executive Briefing")

    write_competitor_analysis(
        ws_competitors, report_data.get("competitors", []), report_date
    )
    print("  Written: Competitor Analysis")

    write_risk_dashboard(
        ws_risk, report_data.get("risk_dashboard", {}), report_date
    )
    print("  Written: Risk Dashboard")

    write_positioning_matrix(
        ws_matrix, report_data.get("positioning_matrix", []), report_date
    )
    print("  Written: Positioning Matrix")

    write_strategic_synthesis(
        ws_synthesis, report_data.get("strategic_synthesis", {}), report_date
    )
    print("  Written: Strategic Synthesis")

    write_raw_data(
        ws_raw, report_data.get("raw_data", ""), report_date
    )
    print("  Written: Raw Data")

    wb.save(filepath)
    print(f"\nReport saved to: {filepath}")
    return filepath
